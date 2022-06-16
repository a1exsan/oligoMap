
from openpyxl import load_workbook
import pandas as pd
import oligoMass.molmassOligo as mmo

class excelSheet():
    """
        Класс взаимодействует со страницей электронной таблицы
    """
    def __init__(self, fn, sheet_name, tamplate={}):
        self.template = tamplate
        wb = load_workbook(filename=fn)
        assert sheet_name in wb.sheetnames, f"sheet {sheet_name} doesn't exist in this document"
        self._sheet = wb[sheet_name]
        self._read_sheet_data()
        self._control_types()

    def _read_sheet_data(self):
        pass

    def _control_types(self):
        for key in self.template.keys():
            for i, value in enumerate(self.sheet[key]):
                if type(value) != self.template[key] and type(value) != type(None):
                    #print(type(value), self.template[key], value)
                    if type(value) == str and self.template[key] == float:
                        self.sheet[key][i] = self.template[key](self.sheet[key][i].replace(',', '.'))
                    elif type(value) == str and self.template[key] == int:
                        fl = float(self.sheet[key][i].replace(',', '.'))
                        self.sheet[key][i] = int(round(fl))
                    else:
                        self.sheet[key][i] = self.template[key](self.sheet[key][i])

    def get_df(self):
        return pd.DataFrame(self.sheet)

    def __call__(self, *args, **kwargs):
        return self.sheet

class RowKeySheet(excelSheet):

    def __init__(self, fn, sheet_name, tamplate={}):
        super().__init__(fn, sheet_name, tamplate)

    def _read_sheet_data(self):
        self.sheet = {}
        for row in self._sheet.values:
            for i, value in enumerate(row):
                if i == 0:
                    self.sheet[value] = []
                    key = value
                else:
                    self.sheet[key].append(value)

class ColKeySheet(excelSheet):

    def __init__(self, fn, sheet_name, tamplate={}):
        super().__init__(fn, sheet_name, tamplate)

    def _read_sheet_data(self):
        self.sheet = {}
        keys = []
        for i, row in enumerate(self._sheet.values):
            for j, value in enumerate(row):
                if i == 0:
                    self.sheet[value] = []
                    keys.append(value)
                else:
                    self.sheet[keys[j]].append(value)


class roadMapSheet():

    def __init__(self):
        self.sheet_name = None
        self.fileName = None
        self.templ = None

class roadMap():
    """
        Класс отображает маршрутную карту синтеза олигонуклеотидов во внутреннее представление
        Маршрутная карта хранится в формате .xlsx
    """
    pass


class synthMethod():
    """
        Класс реализует внутреннее представление метода синтеза олигонуклеотидов
        Вычисляет временные параметры метода, а также расход реагентов
    """

    def __init__(self, fn, sheet_name='synt_programs', column_volume=67):
        super().__init__()
        self.flow_const = 2000 / (60 * 270) # 2 ml/min at 270 rpm
        self.column_volume = column_volume
        self.sheet_name = sheet_name
        self.fileName = fn
        self.templ = {}
        self.templ['Step'] = int
        self.templ['Time'] = float
        self.templ['RPM'] = float
        self.templ['Monitor'] = int
        self.templ['type'] = str
        self.templ['Reagent'] = str
        self.sheet_tab = ColKeySheet(self.fileName, self.sheet_name, self.templ)
        self.data = self.sheet_tab.get_df()[list(self.templ.keys())]
        self.data.dropna(inplace=True)
        self.__calc_param__()

    def __calc_param__(self):
        self.data['volume, ul'] = self.data['Time'] * self.data['RPM'] * self.flow_const
        self.data['CV'] = self.data['volume, ul'] / self.column_volume


class synSimulator():
    def __init__(self, synParam, synMethod, synReagents):
        super().__init__()
        self.synParam = synParam
        self.synMethod = synMethod
        self.synReagents = synReagents
        self.info = None
        self.all_reagents = None
        self.__prepare_param__()

    def get_reagents_mod_subDF(self, base, mod):
        if mod == '':
            return self.synReagents.data[self.synReagents.data['name'] == f'BASE{base}']
        else:
            #print([mod], base)
            #print(self.synReagents.data['name'])
            df = self.synReagents.data[self.synReagents.data['name'].str.contains(f'{mod}', regex=False)]
            #print(df)
            return df[df['name'].str.contains(f'{base}', regex=False)]

    def get_method_mod_subDF(self, base, mod):
        if mod == '':
            return self.synMethod.data[self.synMethod.data['type'] == 'couple']
        else:
            return self.synMethod.data[self.synMethod.data['type'].str.contains(f'{mod}', regex=False)]

    def get_all_bases(self):
        unique_mods = {'name': [], 'mod': [], 'base': []}
        for seq in self.synParam.data['Sequence (5-3)']:
            oligo = mmo.oligoNASequence(seq)
            seqtab = oligo.getSeqTabDF()
            for prefix, base, suffix in zip(seqtab['prefix'], seqtab['nt'], seqtab['suffix']):
                if prefix != '':
                    unique_mods['name'].append(f'{prefix} {base}')
                    unique_mods['mod'].append(f'{prefix}')
                    unique_mods['base'].append(f'_{base}')
                elif suffix != '':
                    unique_mods['name'].append(f'{base} {suffix}')
                    unique_mods['mod'].append(f'{suffix}')
                    unique_mods['base'].append(f'_{base}')
                elif prefix == '' and suffix == '':
                    unique_mods['name'].append(f'{base}')
                    unique_mods['mod'].append(f'')
                    unique_mods['base'].append(f'_{base}')

        self.all_reagents = pd.DataFrame(unique_mods)

    def get_amidites_info(self):
        self.info = {
                'reagent name': [],
                'total reagent volume, ml': [],
                'total reagent mass, g': []
                }
        for reagent in self.all_reagents['name'].unique():
            df = self.all_reagents[self.all_reagents['name'] == reagent]
            couple_num = self.synMethod.data[self.synMethod.data['Reagent'] == 'COUPL'].shape[0]
            sub_method = self.get_method_mod_subDF(base=df['base'].unique()[0], mod=df['mod'].unique()[0])
            sub_reagent = self.get_reagents_mod_subDF(base=df['base'].unique()[0], mod=df['mod'].unique()[0])
            if sub_method.shape[0] > 0 and sub_reagent.shape[0] > 0:
                self.info['reagent name'].append(sub_reagent['name'].unique()[0])
                volume = sub_method[sub_method['Reagent'] == 'BASE']['volume, ul'].sum()\
                     * df.shape[0] * couple_num / 1000
                concentration = sub_reagent[sub_reagent['units'] == 'g']['amount'].max() / \
                            sub_reagent[sub_reagent['units'] == 'ml']['amount'].max()
            #print(concentration, 'g/ml')
                volume += 2000 * 5 / 60
                mass = concentration * volume
                mass = mass + mass * 0.25
                self.info['total reagent volume, ml'].append(mass / concentration)
                self.info['total reagent mass, g'].append(mass)
        self.info['reagent name'].append('ACN amidites')
        self.info['total reagent volume, ml'].append(sum(self.info['total reagent volume, ml']))
        self.info['total reagent mass, g'].append(0)

    def get_reagents_info(self):
        df = self.synMethod.data[self.synMethod.data['type'] == 'main']
        df2 = self.synMethod.data[(self.synMethod.data['type'] == 'removeDMT')|(self.synMethod.data['type'] == 'finish')]
        for m_reag in ['DEBL', 'WASH', 'CAPA', 'CAPB', 'OXID', 'ACTIV']:
            vol = 0.
            for reagent in self.all_reagents['name']:
                vol += df[df['Reagent'] == m_reag]['volume, ul'].sum()
            vol += df2[df2['Reagent'] == m_reag]['volume, ul'].sum()

            mass = 0
            if m_reag == 'ACTIV':
                r_df = self.synReagents.data[self.synReagents.data['name'] == 'ACT']
                concentration = r_df[r_df['units'] == 'g']['amount'].max() / \
                                r_df[r_df['units'] == 'ml']['amount'].max()
                mass = vol * concentration / 1000

            self.info['reagent name'].append(m_reag)
            self.info['total reagent volume, ml'].append(vol / 1000)
            self.info['total reagent mass, g'].append(mass)

    def __prepare_param__(self):
        self.get_all_bases()
        self.get_amidites_info()
        #self.get_reagents_info()

        #print(pd.DataFrame(self.info))



class synthDataBase():
    """
        Класс реализует базу данных методов синтеза
    """
    pass

class Reagents(roadMapSheet):
    """
        Класс реализует внутреннее представление о реагентах для синтеза олигонуклеотидов
        Вычисляет концентрации реагентов, срок годности реагентов
    """
    def __init__(self, fn, sheet_name='Reagents'):
        super().__init__()
        self.sheet_name = sheet_name
        self.fileName = fn
        self.templ = {}
        self.templ['prep_date'] = type(pd.to_datetime('01.01.22'))
        self.templ['name'] = str
        self.templ['substance_name'] = str
        self.templ['amount'] = float
        self.templ['units'] = str
        self.sheet_tab = ColKeySheet(self.fileName, self.sheet_name, self.templ)
        self.data = self.sheet_tab.get_df()[list(self.templ.keys())]
        self.data.dropna(inplace=True)



class synthParams(roadMapSheet):
    """
        Класс реализует внутреннее представление о синтезе олигонуклеотидов
        Вычисляет: время синтеза, расход реагентов
    """
    def __init__(self, fn, sheet_name='synthesis'):
        super().__init__()
        self.fileName = fn
        self.sheet_name = sheet_name
        self.templ = {}
        self.templ['Date'] = type(pd.to_datetime('01.01.22'))
        self.templ['pore_size'] = str
        self.templ['support_amount, mg'] = float
        self.templ['support_capacity, nM/mg'] = float
        self.templ['support_cn'] = str
        self.templ['product_volume, ml'] = float
        self.templ['product_conc, oe/ml'] = float
        self.sheet_tab = RowKeySheet(self.fileName, sheet_name=self.sheet_name, tamplate=self.templ)
        self._reduce_data()
        self.get_calc_params_df()

    def _reduce_data(self):
        self.data = self.sheet_tab.get_df().fillna('null')
        self.data = self.data[self.data['Sequence (5-3)'] != 'null']

    def get_date(self):
        return self.data['Date'].max()

    def get_calc_params_df(self):
        mass, ext_cf, length = [], [], []
        for seq in self.data['Sequence (5-3)']:
            oligos = mmo.oligoNASequence(seq)
            mass.append(oligos.getAvgMass())
            ext_cf.append(oligos.getExtinction())
            length.append(oligos.getSeqLength())

        self.data['Molecular Mass, Da'] = mass
        self.data['Molar extinction, oe*L/mol'] = ext_cf
        self.data['nt length'] = length

        if self.data['product_volume, ml'].max() != 'null' and self.data['product_conc, oe/ml'].max() != 'null':
            self.data['Total amount, OE'] = self.data['product_volume, ml'] * self.data['product_conc, oe/ml']
            self.data['Total amount, nmol'] = self.data['Total amount, OE'] * 1e6 \
                                            / self.data['Molar extinction, oe*L/mol']
            self.data['Max yield, nmol'] = self.data['support_amount, mg'] * self.data['support_capacity, nM/mg']
            self.data['Yield%'] = self.data['Total amount, nmol'] * 100 / self.data['Max yield, nmol']
        else:
            null_list = ['null' for i in range(self.data.shape[0])]
            self.data['Total amount, OE'] = null_list
            self.data['Total amount, nmol'] = null_list
            self.data['Total amount, nmol'] = null_list
            self.data['Max yield, nmol'] = null_list
            self.data['Yield%'] = null_list


class purifMethod(roadMapSheet):
    """
        Класс реализует внутреннее представление о методе очистки олигонуклеотидов
        Вычисляет: время очистки, расход реагентов
    """
    pass

class purifParams():
    """
        Класс реализует внутреннее представление о параметрах очистки олигонуклеотидов
        Вычисляет: выход стадии
    """
    pass

class finalizationProcedure():
    """
        Класс реализует внутреннее представление о финальных этапах подготовки олигонуклеотидов
        Вычисляет: выход время стадии, расход материалов
    """
    pass

def test1():
    wb = load_workbook(filename='/home/alex/Documents/OligoSynt/DB/maps/drive-download-20220511T061617Z-001/synth_250322.xlsx')
    synt_sheet = wb['synthesis']

    for row in synt_sheet.values:
        for value in row:
            print(value, type(value))

    df = pd.DataFrame(synt_sheet.values)
    print(df.info())

def test2():
    es = RowKeySheet('/home/alex/Documents/OligoSynt/DB/maps/drive-download-20220511T061617Z-001/synth_250322.xlsx',
                    sheet_name='synthesis',
                     tamplate={'support_amount, mg': float,
                               'Date': type(pd.to_datetime('01.01.22'))})
    print(es.sheet['support_amount, mg'])
    print(es.sheet['Date'])

    print(pd.DataFrame(es.sheet).dropna().T)

    cs = ColKeySheet('/home/alex/Downloads/synth_080622_LNA_2.xlsx',
                    sheet_name='Reagents',
                     tamplate={'prep_date': type(pd.to_datetime('01.01.22')),
                               'amount': float})
    print(cs.get_df())

def test3():

    sp = synthParams('/home/alex/Downloads/synth_080622_LNA_2.xlsx')
    #print(pd.DataFrame(sp.sheet_tab()).fillna('null').T)
    print(sp.data.T)

def test4():

    #mapName = '/home/alex/Downloads/synth_090622_LNA_3.xlsx'
    mapName = '/home/alex/Downloads/synth_140622_dye_1.xlsx'
    synparams = synthParams(mapName)
    method = synthMethod(mapName)
    reagents = Reagents(mapName)
    simul = synSimulator(synparams, method, reagents)


if __name__=='__main__':
    test4()